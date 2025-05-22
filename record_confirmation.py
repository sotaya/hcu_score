#GPAを算出するプログラム
import re,time,sys
from datetime import datetime
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
import openpyxl
from openpyxl.styles import Alignment
class web:
    def initial(self):
        options=webdriver.ChromeOptions()
        options.add_argument('--headless')
        self.driver=webdriver.Chrome('./chromedriver.exe',options=options)
        self.driver.get('https://ichipol.hiroshima-cu.ac.jp/uniprove_pt/UnLoginAction')
    def transition(self):
        self.driver.find_element(By.XPATH,'//div[@id="contents"]/div[1]/div[1]/ul/li[1]/input[2]').click()
    def decision(self,value):
        if '秀'==value:
            return 4.0
        elif '優'==value:
            return 3.0
        elif '良'==value:
            return 2.0
        elif '可'==value:
            return 1.0
        else:
            return 0.0
    def record(self,UserName,PassWord):
        Enter=int(f'20{UserName[:2]}')
        dates=datetime.now()
        year_now,month_now=dates.year,dates.month
        file_name='report_card.xlsx'
        sum_parameter,sum_total=0.0,0.0
        if month_now<=3:year_now-=1
        self.initial()
        (self.driver.find_element(By.ID,"userid")).send_keys(f"{UserName}")
        (self.driver.find_element(By.ID,"password")).send_keys(f"{PassWord}")
        self.driver.find_element(By.CLASS_NAME,'button_login').click()
        try:
            try:
                self.driver.execute_script('document.querySelector("#retro-student-10002 > div > div > form > div > ul > li:nth-child(11) > a").onclick();')
            except:print('ログインＩＤまたはパスワードが不正です。')
            else:
                time.sleep(2)
                self.driver.switch_to.window(self.driver.window_handles[-1])
                self.driver.find_element(By.XPATH,'//*[@id="idhelp"]/li[3]/a/img').click()
                wait=WebDriverWait(self.driver,10)
                while True:
                    try:
                        wait.until(EC.alert_is_present())
                        Alert(self.driver).accept()
                        time.sleep(1)
                        break
                    except:
                        pass
                time.sleep(2)
                self.driver.find_element(By.XPATH,'//*[@id="academy_menu"]/dl[1]/dd[3]/a').click()
                for year_num in range(Enter,year_now+1):
                    key_years=self.driver.find_element(By.NAME,"txtTacFcy")
                    key_years.clear()
                    key_years.send_keys(f'{year_num}')
                    self.driver.find_element(By.XPATH,'//div[@class="buttonarea"]/input[1]').click()
                    page=self.driver.find_element(By.XPATH,'//div[@id="contents"]/div[1]/div[1]/ul/li[1]')
                    all_page=int((''.join(re.findall(r'[0-9]*',page.text)))[1:])
                    element_num=int((self.driver.find_element(By.XPATH,'//div[@id="contents"]/div[1]/div[1]/ul/li[3]/select/option[1]')).text)
                    total,parameter,data=0.0,0.0,[]
                    head_name=[(self.driver.find_element(By.XPATH,f'//div[@id="contents"]/div[1]/div[2]/table/tbody/tr[1]/th[{i}]')).text for i in range(1,11) if i!=1 and i!=2 and i!=3 and i!=5 and i!=6 and i!=9]
                    for i in range(all_page):
                        try:
                            annual=(self.driver.find_element(By.XPATH,f'//div[@id="contents"]/div[1]/div[2]/table/tbody/tr[2]/td[9]')).text
                            for line in range(element_num):
                                unit=float((self.driver.find_element(By.XPATH,f'//div[@id="contents"]/div[1]/div[2]/table/tbody/tr[{line+2}]/td[8]')).text)
                                evaluation=(self.driver.find_element(By.XPATH,f'//div[@id="contents"]/div[1]/div[2]/table/tbody/tr[{line+2}]/td[10]')).text
                                parameter+=unit
                                total+=(unit*self.decision(evaluation))
                                data.append([(self.driver.find_element(By.XPATH,f'//div[@id="contents"]/div[1]/div[2]/table/tbody/tr[{line+2}]/td[{i}]')).text for i in range(1,11) if i!=1 and i!=2 and i!=3 and i!=5 and i!=6 and i!=9])
                        except:pass
                        else:self.transition()
                    sum_parameter+=parameter
                    sum_total+=total
                    self.driver.find_element(By.XPATH,'//*[@id="contents"]/div[2]/input').click()
                    df=pd.DataFrame(data)
                    df.columns=head_name
                    if year_num==Enter:df.to_excel(file_name,sheet_name=f'{year_num}年度({annual})',index=False)
                    else:
                        with pd.ExcelWriter(file_name,mode='a')as writer:
                            df.to_excel(writer,sheet_name=f'{year_num}年度({annual})',index=False)
                    
                wb=openpyxl.load_workbook(file_name)
                wb_sheet=wb.sheetnames
                for sheet in wb_sheet:
                    ws=wb[f'{sheet}']
                    ws.column_dimensions['B'].width=30
                    row_elements=[ws.cell(row=num,column=1).value for num in range(2,ws.max_row+1)]
                    unanimous=[[row_elements.index(value),row_elements.index(value)+row_elements.count(value)] for value in set(row_elements)]
                    for ind in unanimous:
                        ws.merge_cells(f'A{ind[0]+2}:A{ind[1]+1}')
                        ws[f'A{ind[0]+2}'].alignment=Alignment(horizontal='center',vertical='center')
                    for num in range(2,ws.max_row+1):
                        ws[f'C{num}'].alignment=Alignment(horizontal='center')
                        ws[f'D{num}'].alignment=Alignment(horizontal='center')
                wb.save(file_name)
                print(f'現在、あなたの取得単位数は{sum_parameter}です。')
                print(f'あなたのGPAは{sum_total/sum_parameter}です')
        except:print('GlobalProtectを接続する必要があります')
usenum=input('ユーザー名:')
passnum=input('パスワード:')
if __name__=='__main__':
    web().record(usenum,passnum)
